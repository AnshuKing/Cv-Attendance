def main():
    import numpy as np
    import cv2
    import pickle
    from datetime import datetime, date
    import os.path
    from openpyxl import Workbook
    import openpyxl

    data = {}
    def export2excel(data, src):
        wb = openpyxl.load_workbook(src)
        ws = wb.active
        mr = ws.max_row

        for k,v in data.items():
            mr+=1
            ws.cell(column=1, row=mr, value=k)
            ws.cell(column=3, row=mr, value=v)
        wb.save(src)


    Face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_alt2.xml")
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read("trainner.yml")
    labels={}
    with open("labels.pickle", "rb") as f:
        og_labels = pickle.load(f)
        labels = {v:k for k,v in og_labels.items()}

    cap=cv2.VideoCapture(0)

    while True:
        ret, frame = cap.read()#cv2 converts frame into numpy array by default,
        # so the detectmultiscale func takes an array as (input = gray = frame) as done in faces_train.py
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = Face_cascade.detectMultiScale(gray,1.3, 5)#frame, scalefactor, minneighbord

        for (x,y,w,h) in faces:
            roi_gray = gray[y:y+h, x:x+w]
            roi_color = frame[y:y+h, x:x+w]
            id_, conf = recognizer.predict(roi_gray)
            #print(labels[id_], conf)
            if conf>=40: #Adjust this as u like
                cv2.putText(frame, labels[id_],(x,y),cv2.FONT_HERSHEY_SIMPLEX, 1, (255,0,0),1, cv2.LINE_AA)
                if labels[id_] not in data:
                    data[labels[id_]] = datetime.now().strftime("%H:%M:%S")
                    
##            img_item = "my_img.png"
##            cv2.imwrite(img_item, roi_gray)#saving iamge
            cv2.rectangle(frame, (x,y), (x+w, y+h), (255,0,0), 2)

        cv2.imshow("frame", frame)
        if cv2.waitKey(20) & 0xFF == ord("q"):
            break

    file_name = "Attendances\Attendance_"+(date.today().strftime("%d/%m/%Y"))+'.xlsx'
    base_dir = os.path.dirname(os.path.abspath(__file__))
    att = os.path.join(base_dir, "Attendances")
    file_name = os.path.join(att, 'Attendance_'+(date.today().strftime("%b-%d-%Y"))+'.xlsx')

    if os.path.isfile(file_name):
        export2excel(data, file_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.cell(column = 1, row = 1, value = "Names")
        ws.cell(column = 3, row = 1, value = "Time")
        wb.save(filename = file_name)
        export2excel(data, file_name)

    print(data)
    cap.release()
    cv2.destroyAllWindows()

